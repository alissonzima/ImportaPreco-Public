things = [
    ('1,38', '1.5', 'R$5.439,00'), 
    ('1,84', '1.5', 'R$6.429,00'), 
    ('1,84', '2', 'R$6.539,00'), 
    ('2,3', '2', 'R$7.869,00'), 
    ('2,76', '2', 'R$8.859,00'), 
    ('2,76', '2.5', 'R$8.929,00'), 
    ('2,76', '3', 'R$9.089,00'), 
    ('2,76', '3', 'R$9.679,00'), 
    ('2,76', '3', 'R$10.419,00'), 
    ('3,22', '2.5', 'R$9.919,00'), 
    ('3,22', '3', 'R$10.069,00'), 
    ('3,22', '3', 'R$10.659,00'), 
    ('3,22', '3', 'R$11.409,00'), 
    ('3,68', '3', 'R$11.059,00'), 
    ('3,68', '3', 'R$11.649,00'), 
    ('3,68', '3', 'R$12.389,00'), 
    ('4,14', '3', 'R$12.389,00'), 
    ('4,14', '3', 'R$12.979,00'), 
    ('4,14', '5', 'R$13.479,00'), 
    ('4,14', '3', 'R$13.719,00'), 
    ('4,14', '5', 'R$14.219,00'), 
    ('4,6', '5', 'R$14.459,00'), 
    ('4,6', '5', 'R$15.199,00'), 
    ('5,06', '5', 'R$15.449,00')]

things_length = len(things)

def print_things():
    print('[')
    for i in range(len(things)):
        print(f'({things[i][0]}, {things[i][1]}, {things[i][2]})')
    print(']')

same = 0
kwp = things[0][0]
i = 0

while i < len(things)  :

    old_kwp = kwp

    kwp = things[i][0]
    if old_kwp == kwp:
        same += 1
        if same >= 2:
            placas = float(things[i-1][0].replace(',','.')) * 1000 / 460
            placas_max = float(things[i-1][1]) * 1.4 * 1000 / 460
            if (placas_max - placas) < 2:
                things.pop(i-1)
                i -= 1 
                same -= 1       
            else:
                things.pop(i)
                i -= 1
    else:
        same = 1
    i += 1

print_things()

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Cria Novo workbook
wb = Workbook()
# Seleciona a aba ativa
ws = wb.active
ws["A1"] = "KWP INICIAL"
ws["B1"] = "KWP FINAL"
ws["C1"] = "INVERSOR"
ws["D1"] = "CUSTO KIT"
ws["E1"] = "CUSTO ART"
ws["F1"] = "CUSTO ENG"
ws["G1"] = "GANHO COMERC"
ws["H1"] = "GANHO INSTALAÇÃO"
ws["I1"] = "GANHO INSTALADOR"
ws["J1"] = "ADICIONAL DESP+PROJ"
ws["K1"] = "VALOR VENDA KIT"
ws["L1"] = "SAFELEADS ->"
ws["M1"] = "$ W GERAL"
ws["N1"] = "$ SERVICO"
ws["O1"] = "LAJE ->"
ws["P1"] = "$ LAJE"
ws["Q1"] = "SERV LAJE"
ws["R1"] = "SOLO ->"
ws["S1"] = "$ SOLO"
ws["T1"] = "SERV SOLO"
ws["U1"] = "SEM EST ->"
ws["V1"] = "$ SEM EST"
ws["W1"] = "SERV SEM EST"

for i in range(len(things)):
    kwp = float(things[i][0].replace(",","."))
    ws["A"+str(i+2)] = kwp
    ws["B"+str(i+2)] = kwp
    ws["C"+str(i+2)] = "GROWATT "+things[i][1]
    ws["D"+str(i+2)] = float(things[i][2].replace("R$","").replace(".","").replace(",","."))
    ws["E"+str(i+2)] = float(150)
    if (kwp <= 50):
        ws["F"+str(i+2)] = float(500)
    elif things[i][0] <= 75:
        ws["F"+str(i+2)] = float(700)
    else: 
        ws["F"+str(i+2)] = float(1000)
    if (kwp <= 10.5):
        ws["G"+str(i+2)] = 0.45
    else: 
        ws["G"+str(i+2)] = 0.35
    if (kwp <= 4.49):
        ws["H"+str(i+2)] = kwp * 1000 * 0.3
    elif (kwp <= 10.35):
        ws["H"+str(i+2)] = kwp * 1000 * 0.2
    else: 
        ws["H"+str(i+2)] = kwp * 1000 * 0.15
    if (kwp <= 4.49):
        ws["I"+str(i+2)] = 0.3
    else:
        ws["I"+str(i+2)] = 0.15
    ws["J"+str(i+2)] = 1115
    valor_venda = float(ws["H"+str(i+2)].value) + float(ws["I"+str(i+2)].value) + float(ws["F"+str(i+2)].value) + float(ws["E"+str(i+2)].value) + float(ws["D"+str(i+2)].value) + float(ws["J"+str(i+2)].value) + (float(ws["G"+str(i+2)].value)*float(ws["D"+str(i+2)].value))
    ws["K"+str(i+2)] = valor_venda + valor_venda * float('0.06')
    ws["L"+str(i+2)] = "APP ->"
    ws["M"+str(i+2)] = ws["D"+str(i+2)].value / ws["B"+str(i+2)].value / 1000
    ws["N"+str(i+2)] = ws["K"+str(i+2)].value - ws["M"+str(i+2)].value
    ws["O"+str(i+2)] = "LAJE ->"
    ws["P"+str(i+2)] = ws["M"+str(i+2)].value + 0.15
    novo_custo = ws["P"+str(i+2)].value * ws["A"+str(i+2)].value * 1000
    valor = ws["H"+str(i+2)].value + ws["I"+str(i+2)].value + ws["F"+str(i+2)].value + ws["E"+str(i+2)].value + ws["D"+str(i+2)].value + ws["J"+str(i+2)].value + (ws["G"+str(i+2)].value*novo_custo)
    valor_novo = valor + valor * float('0.06')
    ws["Q"+str(i+2)] = valor_novo - novo_custo
    ws["R"+str(i+2)] = "SOLO ->"
    ws["S"+str(i+2)] = ws["M"+str(i+2)].value + 0.31
    novo_custo = ws["S"+str(i+2)].value * ws["A"+str(i+2)].value * 1000
    valor = ws["H"+str(i+2)].value + ws["I"+str(i+2)].value + ws["F"+str(i+2)].value + ws["E"+str(i+2)].value + ws["D"+str(i+2)].value + ws["J"+str(i+2)].value + (ws["G"+str(i+2)].value*novo_custo)
    valor_novo = valor + valor * float('0.06')
    ws["T"+str(i+2)] = valor_novo - novo_custo
    ws["U"+str(i+2)] = "SEM EST ->"
    ws["V"+str(i+2)] = ws["M"+str(i+2)].value + 0.15
    novo_custo = ws["V"+str(i+2)].value * ws["A"+str(i+2)].value * 1000
    valor = ws["H"+str(i+2)].value + ws["I"+str(i+2)].value + ws["F"+str(i+2)].value + ws["E"+str(i+2)].value + ws["D"+str(i+2)].value + ws["J"+str(i+2)].value + (ws["G"+str(i+2)].value*novo_custo)
    valor_novo = valor + valor * float('0.06')
    ws["W"+str(i+2)] = valor_novo - novo_custo


wb.save('precificacao.xlsx')
