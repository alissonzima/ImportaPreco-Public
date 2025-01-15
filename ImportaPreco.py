''' 
Software for searching photovoltaic kits from the Aldo website, 
according to internal search and dimensioning calculations.

Develop by: Álisson de Moraes Zimermann
Current version: 1.1.5

Patch 1.0.1 
 - Included product id and check digit to excel xls final version.
Patch 1.0.2
 - Included kit description to excel xls final version.
Patch 1.1
 - Change the way it creates excel file. now with 1 big file.
Patch 1.1.1
 - Included area, weight, voltage and phase to excel file.
Patch 1.1.2
 - Ajustado o bug para as placas de 575W.
Patch 1.1.3
 - Ajustado para passar pelo erro 400.
Patch 1.1.4
 - O arquivo agora pode salvar onde quiser e tem pedido de usuário e senha.  
Patch 1.1.5
 - O arquivo trata mais alguns erros no host e ignora alguns produtos.
'''

from threading import Thread
import tkinter as tk
from openpyxl import Workbook
from tkinter import messagebox, filedialog, simpledialog
import datetime
import sys
import os
import requests
import json
import win32console, win32gui
import time
import re

debug = []
# Cria um novo workbook
wb = Workbook()
            
# Seleciona a planilha ativa
ws = wb.active

# Cria um índice de excel para agrupar todas as linhas em um arquivo
excl = 0

# Esconde a janela do DOS
win = win32console.GetConsoleWindow()
win32gui.ShowWindow(win, 0)

# Variável que indica os painéis selecionados
selected = []

# Variável que libera a thread que escreve os arquivos do excel para execução
flag_start_excel = False

# Thread que escreve os arquivos em Excel
thread_excel = None

# Variável global para salvar o caminho do arquivo
file_path = ""

# Índice dos pontos
dot_index = 0

user = ''
password = ''

user = simpledialog.askstring("Usuário", "Insira o nome de usuário")
password = simpledialog.askstring("Senha", "Insira a senha", show='*')

# Cria uma nova janela main
root = tk.Tk()

# Cria uma janela de espera
waiting_window = tk.Tk()
waiting_window.withdraw()

# Cria uma janela final que será mostrada após o final da execução
final_window = tk.Tk()
final_window.withdraw()

def create_window(root: tk, title: str, principal: bool) -> tk:

    """
    Esta função especifica o tamanho e a localização das janelas.

    Args:
        root (tk): a janela que será redimensionada.
        title (str): o título da janela.
        principal (bool): variável que define se a janela é principal ou não.

    """

    # Seletor que define os atributos caso seja janela principal
    if principal:
        root.wm_attributes("-toolwindow", 1)
        root.wm_attributes("-topmost", 1)

    # Variávels que irão pegar o tamanho da tela
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = root.winfo_reqwidth()
    window_height = root.winfo_reqheight()

    # Variáveis que irão fazer os cálculos do tamanho/localização da janela
    x_coordinate = (screen_width / 2) - (window_width / 2)
    y_coordinate = (screen_height / 2) - (window_height / 2)

    # Função que define o tamanho da janela
    root.geometry("+{}+{}".format(int(x_coordinate), int(y_coordinate)))

    # Configura o título da janela
    root.title(title)

    # Retorna a janela configurada
    return root


def open(panels: list) -> None:

    """
    Esta função cria a janela principal para inserir os painéis a serem selecionados.

    Args:
        panels (list): A lista com tuplas de painéis selecionados.

    """

    # Acessa a janela principal
    global root

    # Envia a janela principal para a função create_window
    root = create_window(root, "Buscador Aldo", True)

    # Cria a barra de menu
    menu_bar = tk.Menu(root)
    
    # Cria o "Help" na barra de menu
    help_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="Ajuda", menu=help_menu)

    # Adiciona um "About" para o menu "Help"
    help_menu.add_command(label="Como buscar os preços", command=lambda: tk.messagebox.showinfo("Como buscar os preços", "Esse aplicativo busca os preços automaticamente no site da Aldo.\n\nSelecione quais painéis você quer extrair os valores, depois clique em 'Gerar'"))

    # Configura a barra de menu como a barra principal
    root.config(menu=menu_bar)

    def on_button_click():
    
    '''
        Cria uma função que será chamada quando o botão "gerar" for clicado.

    '''

        global selected, flag_start_excel, root, waiting_window, final_window, file_path
        
        # Abre o dialog para salvar
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        # Inicializa as variáveis selecionadas
        selected = selected

        # Libera a thread que realiza os cálculos e gravações no excel
        flag_start_excel = True

        # Pontos para a espera
        dots = [".", "..", "..."]

        # 
        def update_label_text():
        
        '''
            Função que atualiza o texto da janela de espera.
        
        '''

            # Acessa a janela final
            global final_window, dot_index

            # Checa se a thread_excel ainda está rodando
            if thread_excel.is_alive():

                # Se estiver, mudar o número de pontos ao final
                label.config(text="Aguarde" + dots[dot_index])

                # Se passar de 3, reinicia com 1
                dot_index = (dot_index + 1) % 3

                # Atualiza cada 500 milisegundos
                label.after(500, update_label_text)

            # Se a thread não estiver mais rodando
            else:

                # Mostra a janela final
                final_window.deiconify()

                # Cancela a função update_label_text
                waiting_window.after_cancel(update_label_text)

                # Destrói a janela de espera
                waiting_window.destroy()

        # Envia a janela final para a função que irá redimencioná-la
        final_window = create_window(final_window, "Concluído", False)  

        # Define o texto da janela
        label = tk.Label(final_window, text="Gravação concluída com sucesso.\ n")
        
        # Envia o texto para a janela
        label.pack()
        
        # Cria um botão para finalizar o software e sair do sistema
        end_button = tk.Button(final_window, text="Finalizar", command=sys.exit)
        
        # Envia o botão para a janela
        end_button.pack()
        
        # Envia a janela de espera para a função de formatação
        waiting_window = create_window(waiting_window, "Aguarde mesmo, média atual 3 horas por placa selecionada.", False)
        
        # Insere o texto na janela
        label = tk.Label(waiting_window, text="Aguarde")
        
        # Envia o texto para a janela
        label.pack()
        
        # Exibe a janela
        waiting_window.deiconify()
        
        # Após 500 milissegundos, inicia a função update_label_text
        label.after(500, update_label_text)
        
        # Destrói a janela principal
        root.destroy()

    def on_checkbox_click(index: int) -> None:

      """
      Esta função é chamada quando uma caixa de seleção é clicada.
      Adiciona ou remove o texto da caixa de seleção da lista de opções selecionadas.
      
      Args:
          index (int): O índice da caixa de seleção que foi clicada.
  
      """
  
      # Verifica se a caixa de seleção já está na lista de opções selecionadas
      if index in selected:
          # Se estiver, remove-a
          selected.remove(index)
  
      else:
          # Se não estiver, adiciona-a
          selected.append(index)

    # Cria caixas de seleção com os textos
    for i, panel in enumerate(panels):
        tk.Checkbutton(root, text=panel[0], command=lambda index=i: on_checkbox_click(index), anchor="w").grid(row=i+1, column=0, sticky='w')
    
    # Cria um botão "Gerar"
    generate_button = tk.Button(root, text="Gerar", command=on_button_click, anchor="center", width=7)
    generate_button.grid(row=12, column=0, sticky="w")
    
    # Cria um botão "Fechar"
    close_button = tk.Button(root, text="Fechar", command=root.destroy, anchor="center", width=7)
    close_button.grid(row=12, column=1, sticky='w')
    
    # Inicia a interface gráfica
    root.mainloop()

def fetch_and_calculate(response: str, structure_list: list) -> None:
    """
    Esta função é responsável por pegar a escolha do usuário, 
    montar a lista correta baseada em cálculos internos 
    e entregar a saída em um arquivo Excel, pronto e 
    formatado para inserção em nosso software comercial.

    Args:
        response (str) :
        structure_list (list) :

    """

    structure_list_aux = structure_list.copy()

    # Obtém a variável global excl
    global excl
    excl = excl

    # Separa a string pelos códigos selecionados pelo usuário
    selection = response.split(',')

    # o loop executa até que todos os cálculos com as seleções do usuário sejam concluídos
    while True:
      
            # A resposta sempre recebe a primeira seleção da fila
            response = int(selection[0])
    
            # recebe o nome do painel
            panel = panel_choice[response][0]
    
            # recebe o id do painel
            panel_id = panel_choice[response][1]
    
            structure = structure_list[0][0]
    
            structure_id = structure_list[0][1]
    
            # monta um filtro para enviar à página web com a seleção correta já montada
            filter = str(panel_id) + ';' + str(structure_id) + ';' + str(inverter_id)
    
            # inicializa o array de kits
            kits = []


            # Loop que irá percorrer todos os kits que retornarão com o filtro selecionado pelo usuário
            for x in range(100):
            
                # Envia o filtro para a URL
                tryouts = 0
                while tryouts < 20:
                    try:
                        response = s.post(api_url + 'Produto.svc/getfiltroeprodutosporsegmento',
                                          json=({"slug": "energia-solar", 'origem': 'categoria', 'offset': x, 'filtroAtributos': filter, 'orderby': 1, 'filterId': filter_id, 'idAcionado': '8'}))
                        
                        if response.status_code == 502:
                            raise Exception("Received status code 502")
                        elif response.status_code == 400:
                            break
                        
                        break  # Sai do loop se a requisição POST for bem-sucedida
                                    
                    except Exception as e:
                        tryouts += 1
                        time.sleep(2)     
            
                #print(response)
                #print(response.text)
                #print(response.json())
                # Obtém a resposta
                r = response.json()['Produtos']
                
                # Se existir
                if len(r) > 0:
                    # Adiciona à lista de kits
                    for kit in r:
                        kits.append(kit)
                else: break
            
            products = {}


            # Para cada kit na lista de kits
            for kit in kits:
            
                # Recupera o nome
                name = kit['psg_descricao']
                
                # Verifica kits corretos
                if not any(expression in name for expression in ['ZERO GRID', 'OFF GRID', 'WALLBOX', 'HIBRIDO']):
                
                    # Recupera a descrição
                    description = kit['prd_descricaoInt']
            
                    # Recupera o preço
                    price = kit['prd_preco']
            
                    # Recupera o id do produto e o dígito verificador
                    product_id = '{}-{}'.format(kit['produto_id'], kit['prd_codigodv'])
            
                    tryouts = 0
                    while tryouts < 20:
                        try:
                            # Recupera o HTML da descrição do kit, a cada 2 segundos
                            time.sleep(2)
                            response = s.get(api_url + 'produto.svc/getprodutodisconnected/{}'.format(kit['produto_id']))
                            
                            if response and response.status_code != 500:
                                break
                            else:
                                tryouts += 1    
                                if response.status_code == 500:
                                    time.sleep(600)
                        except Exception as e:
                            tryouts += 1
            
                    if response.status_code == 204:
                        print('erro 204')
                        continue
                    if response.status_code == 400:
                        print('erro 400')
                        continue

                    
                    # Divide o HTML
                    kit_description = response.json()['DescricaoTecnica'].split('</strong><br/>')
                    # print(response.json()['DescricaoTecnica'])
                    weight = re.search(r'Peso.*?(\d+[,.]?\d*)', kit_description[0]).group(1)
                    
                    area_mm = re.search(r'(\d+) *mm *× *(\d+) *mm *× *\d+ *mm', kit_description[0])
                    if area_mm:
                        l = int(area_mm.group(1)) / 1000
                        w = int(area_mm.group(2)) / 1000
                        area = w * l
                    else:
                        area = '2,5'
                    for row in kit_description:
                        
                        match_voltage = re.search(r'(\d+)V', row)
                        if match_voltage:
                            voltage = match_voltage.group()
                        
                        match_phase = re.search(r'\b(\w+FASICO|TRIF)\b', row)
                        if match_phase:
                            phase = match_phase.group()
                    
                    # Variável que indicará a próxima linha
                    found = False
                    
                    # Variável que armazenará a próxima linha
                    next_line = ""
                    
                    # Loop que busca por 'composto por' no HTML
                    for row in kit_description:
                    
                        # Se a linha já for encontrada, armazena o conteúdo e quebra
                        if found:
                            next_line = row
                            break
                    
                        # Se 'composto por' estiver na linha, define o flag como verdadeiro
                        if 'composto por' in row:
                            found = True
                    
                    # Divide next_line para obter apenas a primeira linha
                    kit_description = next_line.split('<img title')
                    
                    # Obtém a linha correta da descrição do kit
                    kit_description = kit_description[0]
                    
                    # Linhas de depuração
                    #global debug
                    #debug.append([name, description, product_id, kit_description[0]])
                    
                    # Realiza o trabalho no excel, baseado nos resultados recuperados
                    matriz = description.split(' ')

                    kwp = None
                    inv = None
                    lista = [None] * 9  # Inicializa a lista com 9 elementos None

                    for row in matriz:
                        row = row + ' '
                        if 'KWP' in row:
                            idx_placa = row.find('K')
                            kwp = row[:idx_placa]
                            lista = [kwp, inv, price, kit_description, product_id, area, weight, phase, voltage]
                        elif 'KW ' in row:
                            idx_inv = row.find('K')
                            inv = row[:idx_inv]
                            lista[1] = inv  # Atualiza o valor de inv na lista

                        # Adiciona a lista à products kwp
                        if kwp is not None and lista is not None and inv is not None:
                            if kwp not in products:
                                products[kwp] = []
                            products[kwp].append(lista)
                            break

            # Ordena cada lista de products no dicionário
            for kwp in products:
                products[kwp] = sorted(products[kwp], key = lambda item : float(item[0].replace(',','.')))

            # Seleciona o kit apropriado para cada kwp
            selected_kits = []
            for kwp in products:
                if float(kwp.replace(',', '.')) <= 10:  # Para sistemas abaixo de 10 kwp
                    for i in range(len(products[kwp])):
                        placas = float(products[kwp][i][0].replace(',','.')) * 1000 / 460
                        placas_max = float(products[kwp][i][1]) * 1.4 * 1000 / 460
                        if (placas_max - placas) >= 2 or len(products[kwp]) == 1:
                            selected_kits.append(products[kwp][i])
                            break
                else:  # Para sistemas acima de 10 kwp
                    for i in range(len(products[kwp])):
                        if products[kwp][i][8] == '380V':
                            selected_kits.append(products[kwp][i])
                            break
                    else:  # Se não houver kit 380V
                        for i in range(len(products[kwp])):
                            placas = float(products[kwp][i][0].replace(',','.')) * 1000 / 460
                            placas_max = float(products[kwp][i][1]) * 1.4 * 1000 / 460
                            if (placas_max - placas) >= 2 or len(products[kwp]) == 1:
                                selected_kits.append(products[kwp][i])
                                break

            # Agora selected_kits contém o kit selecionado para cada kwp
            things = selected_kits

            def create_workbook() :

                global wb, ws

                wb=wb
                ws=ws

                # Colunas do Excel
                ws["A1"] = "PotenciaCC"
                ws["B1"] = "QtdInversores"
                ws["C1"] = "MarcaInversor"
                ws["D1"] = "PrecoCusto"
                ws["E1"] = "CUSTO ART"
                ws["F1"] = "CUSTO ENG"
                ws["G1"] = "GANHO COMERC"
                ws["H1"] = "GANHO INSTALAÇÃO"
                ws["I1"] = "GANHO INSTALADOR"
                ws["J1"] = "ADICIONAL DESP+PROJ"
                ws["K1"] = "PrecoVenda"
                ws["L1"] = "ItensKit"
                ws["M1"] = "QtdComponente1"
                ws["N1"] = "NomeComponente1"
                ws["O1"] = "QtdComponente2"
                ws["P1"] = "NomeComponente2"
                ws["Q1"] = "QtdComponente3"
                ws["R1"] = "NomeComponente3"
                ws["S1"] = "QtdComponente4"
                ws["T1"] = "NomeComponente4"
                ws["U1"] = "QtdComponente5"
                ws["V1"] = "NomeComponente5"
                ws["X1"] = "TipoTelhado"
                ws["Y1"] = "MarcaModulo"
                ws["Z1"] = "QtdModulos"
                ws["AA1"] = "PotenciaModulo"
                ws["AB1"] = "Descricao"
                ws["AC1"] = "QtdStringBox"
                ws["AD1"] = "StringBox"
                ws["AE1"] = "Inversor"
                ws["AF1"] = "Modulo"
                ws["AG1"] = "PotenciaInversor"
                ws["AH1"] = "AreaModulo"
                ws["AI1"] = "PesoModulo"
                ws["AJ1"] = "Fase"
                ws["AK1"] = "Tensao"

            if excl == 0:
                create_workbook()

            # Cálculos
            for i in range(len(things)):

                global wb, ws

                wb=wb
                ws=ws            

                kwp = float(things[i][0].replace(",","."))
                ws["A"+str(excl+2)] = kwp
                ws["C"+str(excl+2)] = 'GROWATT'
                ws["D"+str(excl+2)] = things[i][2]
                ws["E"+str(excl+2)] = float(150)
                if (kwp <= 50):
                    ws["F"+str(excl+2)] = float(500)
                elif kwp <= 75:
                    ws["F"+str(excl+2)] = float(700)
                else: 
                    ws["F"+str(excl+2)] = float(1000)
                if (kwp <= 10.5):
                    ws["G"+str(excl+2)] = 0.45
                else: 
                    ws["G"+str(excl+2)] = 0.35
                if (kwp <= 4.49):
                    ws["H"+str(excl+2)] = kwp * 1000 * 0.3
                elif (kwp <= 10.35):
                    ws["H"+str(excl+2)] = kwp * 1000 * 0.2
                else: 
                    ws["H"+str(excl+2)] = kwp * 1000 * 0.15
                if (kwp <= 4.49):
                    ws["I"+str(excl+2)] = kwp * 1000 * 0.3
                else:
                    ws["I"+str(excl+2)] = kwp * 1000 * 0.15
                ws["J"+str(excl+2)] = 1115
                valor_venda = float(ws["H"+str(excl+2)].value) + float(ws["I"+str(excl+2)].value) + float(ws["F"+str(excl+2)].value) + float(ws["E"+str(excl+2)].value) + float(ws["D"+str(excl+2)].value) + float(ws["J"+str(excl+2)].value) + (float(ws["G"+str(excl+2)].value)*float(ws["D"+str(excl+2)].value))
                ws["K"+str(excl+2)] = valor_venda + valor_venda * float('0.06')
                ws["L"+str(excl+2)] = things[i][3]
                ws["AB"+str(excl+2)] = 'CÓDIGO ALDO: ' + things[i][4]
                ws["X"+str(excl+2)] = structure
                ws["AH"+str(excl+2)] = things[i][5]
                ws["AI"+str(excl+2)] = things[i][6].replace(',','.')
                ws["AJ"+str(excl+2)] = things[i][7]
                ws["AK"+str(excl+2)] = things[i][8]
                ws["Y"+str(excl+2)] = panel
                
                potencia_modulo = panel.split(' ')
                
                for linha in potencia_modulo :
                    if not linha.isalpha() :        
                        ws["AA"+str(excl+2)] = "".join(re.findall(r'\d+', linha))
                
                itens = things[i][3]

                itens_separados = re.split('<br[\/]?>',itens)
                index = 0
                for item in itens_separados:
                    linha = item.strip().split(" ",1)
                    if len(linha) > 1:
                        qtd = linha[0]
                        desc = linha[1]                    
                        if 'MICRO' not in item and 'INVERSOR' in item:
                            ws["B"+str(excl+2)] = qtd
                            ws["AE"+str(excl+2)] = desc
                            potencia_inversor = re.search(r'(\d+(\.\d+)?)KW', item).group(1)
                            ws["AG"+str(excl+2)] = potencia_inversor
                        elif 'PAINEL' in item:
                            ws["Z"+str(excl+2)] = qtd
                            ws["AF"+str(excl+2)] = desc   
                        elif 'STRING BOX' in item:
                            ws["AC"+str(excl+2)] = qtd
                            ws["AD"+str(excl+2)] = desc
                        elif any(expression in item for expression in ['CABO', 'STAUBLI', 'ESTRUTURA']):                     
                            letras = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
                            ws[letras[index]+str(excl+2)] = qtd
                            ws[letras[index+1]+str(excl+2)] = desc
                            index += 2
                excl += 1

            # Obtém a data atual
            current_date = datetime.datetime.now()
            
            # Formata a data como dia-mês-ano
            formatted_date = current_date.strftime("%d-%m-%Y")
            
            # Função para salvar o arquivo ao final da execução
            def save_file():
                
                global file_path
                
                # Verifica se um caminho de arquivo foi selecionado
                if file_path:
                    # Salva o arquivo no local selecionado
                    wb.save(file_path)
            
            # Remove a estrutura da lista
            structure_list.pop(0)
            
            if not structure_list:
                # Preenche novamente a lista de estruturas
                structure_list = structure_list_aux.copy()
            
                # Remove o painel da lista
                selection.pop(0)
            
                # Se não houver mais painéis
                if not selection:
                    # Salva o arquivo
                    save_file()
            
                    # Sai da função
                    break


def start_thread(structure_list: list) -> None:

    '''
    Função que inicia a thread do Excel, 
    enquanto a outra continuará atualizando a janela principal.

    Args:
        structure_list (list): A lista de tuplas das estruturas selecionadas
    '''

    global flag_start_excel, thread_excel
    # Permanece no loop até que a flag para iniciar os cálculos do Excel seja definida como verdadeira
    while not flag_start_excel:
        pass
    # Cria uma string com as seleções do usuário
    response = ", ".join(str(x) for x in selected)
    # Inicia uma nova thread para os cálculos do Excel, passando a string criada como parâmetro
    thread_excel = Thread(target=fetch_and_calculate, args=[response, structure_list])
    # Inicia a thread
    thread_excel.start()



if __name__ == "__main__":

    # Iniciando o arquivo principal criando uma sessão
    with requests.Session() as s :

        # URL da API da Aldo
        api_url = 'https://www.aldo.com.br/wcf/'

        # Enviando login e headers para a API
        response = s.get('https://www.aldo.com.br/login', allow_redirects=False)
        s.headers.update({'Content-Type': 'application/json'})
        s.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 OPR/92.0.0.0'})
        response = s.post(api_url + 'Login.svc/login', json=({'usuario': user, 'senha': password}), headers=s.headers)
        
        filter_id = 0
        # Obtendo itens da Aldo de acordo com a categoria de energia solar
        response = s.post(api_url + 'Produto.svc/getfiltroeprodutosporsegmento', json=({'slug': 'energia-solar', 'origem': 'categoria', "filtroAtributos": None, "offset": 0, "filterId": filter_id, "orderby": 1}))

        # Inicializa array de painéis
        panel_choice = []

        structure_list = []

        # Loop para obter dados de estruturas, inversores e painéis do response
        for dado in json.loads(response.content)['Filtros']['Dados'] :

            if dado['Descricao'] == 'ESTRUTURA':
                estruturas = dado['Valores']
            elif dado['Descricao'] == 'INVERSOR SOLAR':
                inversores = dado['Valores']
            elif dado['Descricao'] == 'PAINEL SOLAR':
                paineis = dado['Valores']
        
        # Loop para obter dados específicos de estrutura
        for estrutura in estruturas:
            if estrutura['Descricao'] in ["PARAFUSO ESTRUTURAL MADEIRA", "TELHA COLONIAL GANCHO", "TELHA METALICA PERFIL 55CM", "SOLO", "SEM ESTRUTURA", "LAJE TRIANGULO"] :
                structure_list.append((estrutura['Descricao'],estrutura['Id']))

        # Loop para obter dados específicos de inversor
        for inversor in inversores:
            if inversor['Descricao'] == 'GROWATT':
                inverter_id = inversor['Id']

        # Loop para obter dados específicos de painel
        for painel in paineis:
            panel_choice.append((painel['Descricao'],painel['Id']))

        # Inicializa a thread de cálculo interna
        thread_interna = Thread(target=start_thread, group=None, args=[structure_list])
        thread_interna.start()

        # Chama a função 'open' para criar a tela principal
        open(panel_choice)
